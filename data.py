import streamlit as st
import yfinance as yf
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

st.set_page_config(page_title="📈 Stock Data Downloader", layout="wide")

st.title("📈 Yahoo Finance Stock Downloader")
st.markdown("Download historical stock data for selected companies into a clean Excel sheet.")

# 🎯 Predefined list of NSE tickers (sample list for demo)
tickers_list = sorted([
    "ASIANPAINT.NS", "TATAMOTORS.NS", "RELIANCE.NS", "HDFCBANK.NS", "ITC.NS",
    "INFY.NS", "SBIN.NS", "ADANIENT.NS", "ULTRACEMCO.NS", "LT.NS"
])

# --- Sidebar Controls ---
with st.sidebar:
    st.header("⚙️ Download Settings")

    selected_tickers = st.multiselect("Choose Tickers", tickers_list)
    period = st.selectbox("Select Period", [
        "1 day", "5 days", "1 month", "3 months", "6 months", "1 year", "2 years", "5 years", "10 years", "YTD", "Max"
    ])
    interval = st.selectbox("Select Interval", ["Daily", "Weekly", "Monthly"])
    download_btn = st.button("📥 Download Excel")

# Mapping to Yahoo Finance codes
period_map = {
    "1 day": "1d", "5 days": "5d", "1 month": "1mo", "3 months": "3mo",
    "6 months": "6mo", "1 year": "1y", "2 years": "2y", "5 years": "5y",
    "10 years": "10y", "YTD": "ytd", "Max": "max"
}
interval_map = {"Daily": "1d", "Weekly": "1wk", "Monthly": "1mo"}

# --- Download logic ---
def fetch_stock_data(tickers, period, interval):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine="openpyxl")

    for ticker in tickers:
        try:
            df = yf.download(ticker, period=period, interval=interval, progress=False)
            if not df.empty:
                df.to_excel(writer, sheet_name=ticker[:31])
        except Exception as e:
            st.error(f"❌ Error downloading {ticker}: {e}")
    writer.close()

    # Auto-adjust column width
    wb = openpyxl.load_workbook(output)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    new_output = BytesIO()
    wb.save(new_output)
    return new_output

# --- Trigger Download ---
if download_btn:
    if not selected_tickers:
        st.warning("Please select at least one ticker from the sidebar.")
    else:
        with st.spinner("Fetching data..."):
            excel_file = fetch_stock_data(
                selected_tickers,
                period=period_map[period],
                interval=interval_map[interval]
            )
        st.success("✅ Download Ready!")

        st.download_button(
            label="📂 Download Excel File",
            data=excel_file.getvalue(),
            file_name="StockData.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
