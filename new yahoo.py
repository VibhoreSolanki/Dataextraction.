# Final Version: Stock Data Downloader with Excel Export, Multiselect, and Smart UI

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
    from ttkwidgets.autocomplete import AutocompleteCombobox
    import yfinance as yf
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    import os
except ModuleNotFoundError as e:
    print("\nERROR:", e)
    print("This application requires 'tkinter' and other GUI-related modules to run.")
    print("Please make sure you're running this code in a local Python environment where 'tkinter' is installed.")
    exit()

class StockDataDownloaderApp:
    def __init__(self, master):
        self.master = master
        master.title("Yahoo Finance Data Downloader")
        master.geometry("700x600")
        master.configure(bg="#f0f4f8")
        master.resizable(False, False)

        # Styling
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", background="#f0f4f8", font=("Segoe UI", 10))
        style.configure("TButton", font=("Segoe UI", 10), background="#1e88e5", foreground="white")
        style.map("TButton", background=[("active", "#1565c0")])

        # --- Variables ---
        self.search_query = tk.StringVar()
        self.selected_ticker_info = tk.StringVar(value="No tickers selected")
        self.selected_tickers_display = tk.StringVar(value="None")
        self.period_var = tk.StringVar(value="1 year")
        self.interval_var = tk.StringVar(value="Weekly")
        self.status_text = tk.StringVar()
        self.search_results_map = {}

        # Dummy ticker list (replace with your full list if needed)
        self.all_tickers = [ "ASIANPAINT.NS", "TATAMOTORS.NS", "RELIANCE.NS", "HDFCBANK.NS", "ITC.NS",
    "INFY.NS", "SBIN.NS", "ADANIENT.NS", "ULTRACEMCO.NS", "LT.NS",
    "WIPRO.NS", "TECHM.NS", "MARUTI.NS", "DMART.NS", "COALINDIA.NS",
    "ONGC.NS", "JSWSTEEL.NS", "HCLTECH.NS", "AXISBANK.NS", "BAJAJ-AUTO.NS",
    "ICICIBANK.NS", "DIVISLAB.NS", "SUNPHARMA.NS", "BHARTIARTL.NS", "KOTAKBANK.NS",
    "POWERGRID.NS", "TCS.NS", "NTPC.NS", "HINDUNILVR.NS", "HINDALCO.NS",
    "BPCL.NS", "BAJAJFINSV.NS", "BAJFINANCE.NS", "GRASIM.NS", "INDUSINDBK.NS",
    "CIPLA.NS", "ADANIPORTS.NS", "EICHERMOT.NS", "BRITANNIA.NS", "TITAN.NS",
    "HEROMOTOCO.NS", "SHREECEM.NS", "UPL.NS", "TATACONSUM.NS", "TATASTEEL.NS",
    "HDFCLIFE.NS", "SBILIFE.NS", "BAJAJHLDNG.NS", "M&M.NS", "NESTLEIND.NS",
    "PIDILITIND.NS", "DRREDDY.NS", "AMBUJACEM.NS", "ABB.NS", "ATGL.NS",
    "BANDHANBNK.NS", "BERGEPAINT.NS", "BEL.NS", "CANBK.NS", "CHOLAFIN.NS",
    "DABUR.NS", "GAIL.NS", "GODREJCP.NS", "HAVELLS.NS", "ICICIPRULI.NS",
    "IDFCFIRSTB.NS", "IOC.NS", "JINDALSTEL.NS", "LUPIN.NS", "MUTHOOTFIN.NS",
    "OBEROIRLTY.NS", "PNB.NS", "RECLTD.NS", "SRF.NS", "TRENT.NS",
    "TVSMOTOR.NS", "UNIONBANK.NS", "VOLTAS.NS", "ZOMATO.NS", "NAVINFLUOR.NS",
    "TIINDIA.NS", "ESCORTS.NS", "GUJGASLTD.NS", "GLENMARK.NS", "IGL.NS",
    "AUBANK.NS", "ALKEM.NS", "DALBHARAT.NS", "DEEPAKNTR.NS", "CUMMINSIND.NS",
    "FORTIS.NS", "GMRINFRA.NS", "IDBI.NS", "IRCTC.NS", "JSWENERGY.NS",
    "MINDTREE.NS", "NMDC.NS", "PERSISTENT.NS", "POLYCAB.NS", "RAJESHEXPO.NS",
    "SAIL.NS", "TATAPOWER.NS", "UCOBANK.NS", "UJJIVANSFB.NS", "YESBANK.NS",
    "ZYDUSLIFE.NS", "BALKRISIND.NS", "BATAINDIA.NS", "BANKINDIA.NS", "BLUEDART.NS",
    "BSOFT.NS", "CASTROLIND.NS", "CONCOR.NS", "CREDITACC.NS", "CROMPTON.NS",
    "DCBBANK.NS", "EXIDEIND.NS", "FEDERALBNK.NS", "FINEORG.NS", "GODREJPROP.NS",
    "HATSUN.NS", "HFCL.NS", "HINDCOPPER.NS", "IIFL.NS", "INDIGO.NS",
    "INDOSTAR.NS", "JUBLFOOD.NS", "LALPATHLAB.NS", "LINDEINDIA.NS", "MAHINDCIE.NS",
    "METROPOLIS.NS", "MOTILALOFS.NS", "MRPL.NS", "NATIONALUM.NS", "NAUKRI.NS",
    "PGHL.NS", "PRAJIND.NS", "PRINCEPIPE.NS", "RBLBANK.NS", "RENUKA.NS",
    "SBICARD.NS", "SHRIRAMFIN.NS", "SPARC.NS", "STLTECH.NS", "SYNGENE.NS",
    "TATACHEM.NS", "TCIEXP.NS", "TRIDENT.NS", "TTKPRESTIG.NS", "VGUARD.NS",
    "VINATIORGA.NS", "WHIRLPOOL.NS", "ZEEL.NS", "CENTURYTEX.NS", "CESC.NS",
    "EIDPARRY.NS", "ENGINERSIN.NS", "FSL.NS", "GESHIP.NS", "GODFRYPHLP.NS",
    "HINDPETRO.NS", "INDIACEM.NS", "IRFC.NS", "ISEC.NS", "JUBLINGREA.NS",
    "KEI.NS", "MAHABANK.NS", "MANGCHEFER.NS", "MCX.NS", "NETWORK18.NS",
    "NHPC.NS", "OIL.NS", "PETRONET.NS", "PNBHOUSING.NS", "RITES.NS",
    "RVNL.NS", "SJVN.NS", "TANLA.NS", "TATAELXSI.NS", "THYROCARE.NS",
    "TORNTPOWER.NS", "UJJIVAN.NS", "VAKRANGEE.NS", "WELCORP.NS", "WELSPUNIND.NS"
    "AARTIDRUGS.NS", "AARTIIND.NS", "ADANIGREEN.NS", "ADANITRANS.NS", "AFFLE.NS",
    "AJANTPHARM.NS", "ALKYLAMINE.NS", "AMARAJABAT.NS", "ANURAS.NS", "APLAPOLLO.NS",
    "APOLLOHOSP.NS", "APOLLOTYRE.NS", "ARVINDFASN.NS", "ASAHIINDIA.NS", "ASHOKLEY.NS",
    "ASTRAZEN.NS", "ASTRAL.NS", "ATUL.NS", "AUROPHARMA.NS", "AVANTIFEED.NS",
    "BALAMINES.NS", "BALRAMCHIN.NS", "BASF.NS", "BAYERCROP.NS", "BEML.NS",
    "BHEL.NS", "BIOCON.NS", "BIRLACORPN.NS", "BODALCHEM.NS", "BOMDYEING.NS",
    "BORORENEW.NS", "BRIGADE.NS", "CADILAHC.NS", "CAPLIPOINT.NS", "CARBORUNIV.NS",
    "CEATLTD.NS", "CENTRALBK.NS", "CENTURYPLY.NS", "CERA.NS", "CGPOWER.NS",
    "CHEMPLASTS.NS", "CHENNPETRO.NS", "COCHINSHIP.NS", "COFORGE.NS", "COLPAL.NS",
    "COROMANDEL.NS", "CUB.NS", "CYIENT.NS", "DBCORP.NS", "DCMSHRIRAM.NS",
    "DELTACORP.NS", "DEVYANI.NS", "DHANI.NS", "DISHTV.NS", "DIXON.NS",
    "ECLERX.NS", "EDELWEISS.NS", "EICHERMOT.NS", "EMAMILTD.NS", "ENDURANCE.NS",
    "EPL.NS", "ERIS.NS", "EVEREADY.NS", "EVERESTIND.NS", "FACT.NS",
    "FCONSUMER.NS", "FINCABLES.NS", "FINEOTEX.NS", "FLUOROCHEM.NS", "FMGOETZE.NS",
    "GEPIL.NS", "GHCL.NS", "GICRE.NS", "GILLETTE.NS", "GLAXO.NS",
    "GNFC.NS", "GODREJAGRO.NS", "GPIL.NS", "GREAVESCOT.NS", "GREENPLY.NS",
    "GRINDWELL.NS", "GSFC.NS", "GSKCONS.NS", "GULFOILLUB.NS", "HCG.NS",
    "HEG.NS", "HEIDELBERG.NS", "HERITGFOOD.NS", "HFCL.NS", "HIKAL.NS",
    "HIMATSEIDE.NS", "HINDZINC.NS", "HSCL.NS", "IEX.NS", "IFBIND.NS",
    "IGARASHI.NS", "IIFLWAM.NS", "IL&FSTRANS.NS", "IMFA.NS", "INDHOTEL.NS",
    "INDIAMART.NS", "INDIGOPNTS.NS", "INDOSTAR.NS", "INDTERRAIN.NS", "INSECTICID.NS",
    "INTELLECT.NS", "IRB.NS", "ISGEC.NS", "ITI.NS", "JAGRAN.NS",
    "JAICORPLTD.NS", "JAMNAAUTO.NS", "JBCHEPHARM.NS", "JCHAC.NS", "JINDALPOLY.NS",
    "JKCEMENT.NS", "JKLAKSHMI.NS", "JKPAPER.NS", "JKTYRE.NS", "JSL.NS",
    "JTEKTINDIA.NS", "JUBLINGREA.NS", "JYOTHYLAB.NS", "KALPATPOWR.NS", "KANSAINER.NS",
    "KARURVYSYA.NS", "KCP.NS", "KDDL.NS", "KEC.NS", "KIRLOSENG.NS",
    "KNRCON.NS", "KPRMILL.NS", "KRBL.NS", "LAXMIMACH.NS", "LEMONTREE.NS",
    "LINCOLN.NS", "LTI.NS", "LTTS.NS", "LUXIND.NS", "MAHINDCIE.NS"
    "MAHSCOOTER.NS", "MANAPPURAM.NS", "MARKSANS.NS", "MASTEK.NS", "MAZDOCK.NS",
    "MEGH.NS", "MINDACORP.NS", "MOIL.NS", "MPHASIS.NS", "MRF.NS",
    "MSTCLTD.NS", "NATCOPHARM.NS", "NBCC.NS", "NBVENTURES.NS", "NCC.NS",
    "NEULANDLAB.NS", "NFL.NS", "NH.NS", "NILKAMAL.NS", "NRBBEARING.NS",
    "NSLNISP.NS", "NUCLEUS.NS", "OLECTRA.NS", "ORIENTCEM.NS", "ORIENTELEC.NS",
    "PAGEIND.NS", "PCBL.NS", "PFIZER.NS", "PHOENIXLTD.NS", "PIIND.NS",
    "PILITA.NS", "PILANIINV.NS", "PRAKASH.NS", "PRSMJOHNSN.NS", "PTC.NS",
    "QUESS.NS", "RADICO.NS", "RAIN.NS", "RAMCOCEM.NS", "RALLIS.NS",
    "RATNAMANI.NS", "RCF.NS", "REDINGTON.NS", "RELAXO.NS", "REPCOHOME.NS",
    "RICOAUTO.NS", "ROUTE.NS", "RTNINDIA.NS", "RVNL.NS", "SADBHAV.NS",
    "SANSERA.NS", "SANOFI.NS", "SASTASUNDR.NS", "SBC.NS", "SCHAEFFLER.NS",
    "SCI.NS", "SEQUENT.NS", "SESHAPAPER.NS", "SHK.NS", "SHOPERSTOP.NS",
    "SHYAMMETL.NS", "SIS.NS", "SJS.NS", "SKFINDIA.NS", "SOLARA.NS",
    "SONATSOFTW.NS", "SOUTHBANK.NS", "SPANDANA.NS", "SPENCERS.NS", "SPICEJET.NS",
    "SPTL.NS", "STAR.NS", "STEL.NS", "STERTOOLS.NS", "SUBEXLTD.NS",
    "SUDARSCHEM.NS", "SUVENPHAR.NS", "SWANENERGY.NS", "TASTYBITE.NS", "TCNSBRANDS.NS",
    "TCPLPACK.NS", "TEAMLEASE.NS", "TEGA.NS", "TEXRAIL.NS", "THEINVEST.NS",
    "THERMAX.NS", "THOMASCOOK.NS", "THYROCARE.NS", "TIDEWATER.NS", "TIMETECHNO.NS",
    "TIRUMALCHM.NS", "TITAGARH.NS", "TNPL.NS", "TOKYOPLAST.NS", "TORNTPHARM.NS",
    "TPLPLASTEH.NS", "TRANSCHEM.NS", "TRF.NS", "TRIVENI.NS", "TTML.NS",
    "TV18BRDCST.NS", "TVTODAY.NS", "UBL.NS", "UCALFUEL.NS", "UFLEX.NS",
    "UGARSUGAR.NS", "UJJIVANSFB.NS", "UMAEXPORTS.NS", "UNIDT.NS", "UNIVCABLES.NS",
    "UTIAMC.NS", "VAIBHAVGBL.NS", "VARROC.NS", "VASWANI.NS", "VBL.NS",
    "VENKEYS.NS", "VETO.NS", "VGUARD.NS", "VISHNU.NS", "VISHWARAJ.NS",
    "VMART.NS", "VOLTAMP.NS", "VRLLOG.NS", "VSSL.NS", "WABAG.NS",
    "WABCOINDIA.NS", "WANBURY.NS", "WATERBASE.NS", "WEIZMANIND.NS", "WELENT.NS",
    "WESTLIFE.NS", "WHEELS.NS", "WIPRO.NS", "WONDERLA.NS", "WSTCSTPAPR.NS",
    "XCHANGING.NS", "YESBANK.NS", "ZEELEARN.NS", "ZEEMEDIA.NS", "ZENSARTECH.NS",
    "ZODIACLOTH.NS", "ZYDUSWELL.NS", "21STCENMGM.NS", "3IINFOTECH.NS", "3MINDIA.NS",
    "63MOONS.NS", "A2ZINFRA.NS", "AAKASH.NS", "AARON.NS", "AARTISURF.NS",
    "AAVAS.NS", "ABAN.NS", "ABBOTINDIA.NS", "ABCAPITAL.NS", "ABFRL.NS",
    "ACC.NS", "ACCELYA.NS", "ACCURACY.NS", "ACE.NS", "ADFFOODS.NS",
    "ADORWELD.NS", "ADVANIHOTR.NS", "AEGISCHEM.NS", "AFFLE.NS", "AGARIND.NS"
    "AGRITECH.NS", "AHLADA.NS", "AHLEAST.NS", "AHLUCONT.NS", "AIAENG.NS",
    "AIRAN.NS", "AJRINFRA.NS", "AKASH.NS", "AKZOINDIA.NS", "ALANKIT.NS",
    "ALBERTDAVD.NS", "ALCHEM.NS", "ALEMBICLTD.NS", "ALICON.NS", "ALKALI.NS",
    "ALLCARGO.NS", "ALLSEC.NS", "ALOKINDS.NS", "ALPHAGEO.NS", "AMBER.NS",
    "AMJLAND.NS", "ANDHRAPAP.NS", "ANDHRSUGAR.NS", "ANGELONE.NS", "ANIKINDS.NS",
    "ANSALAPI.NS", "ANTGRAPHIC.NS", "APCOTEXIND.NS", "APEX.NS", "APLAPOLLO.NS",
    "APLLTD.NS", "APOLLO.NS", "APOLLOPIPE.NS", "APTUS.NS", "ARCHIDPLY.NS",
    "ARENTERP.NS", "ARIES.NS", "ARSSINFRA.NS", "ARTEMISMED.NS", "ARTNIRMAN.NS",
    "ARVEE.NS", "ARVIND.NS", "ASAHISONG.NS", "ASAL.NS", "ASHAPURMIN.NS",
    "ASHIMASYN.NS", "ASIANENE.NS", "ASIANHOTNR.NS", "ASPINWALL.NS", "ASTEC.NS",
    "ASTRAMICRO.NS", "ATHARVENT.NS", "ATLANTA.NS", "ATULAUTO.NS", "AUBANK.NS",
    "AURIONPRO.NS", "AUTOAXLES.NS", "AUTOIND.NS", "AVADHSUGAR.NS", "AVANTEL.NS",
    "AXITA.NS", "AYMSYNTEX.NS", "BAFNAPH.NS", "BAGFILMS.NS", "BAJAJCON.NS",
    "BAJAJELEC.NS", "BAJAJHCARE.NS", "BALAJITELE.NS", "BALPHARMA.NS", "BANG.NS",
    "BANKBARODA.NS", "BANSWRAS.NS", "BARBEQUE.NS", "BASF.NS", "BATAINDIA.NS",
    "BAYERCROP.NS", "BBL.NS", "BCLIND.NS", "BCP.NS", "BEML.NS",
    "BENGALASM.NS", "BERGEPAINT.NS", "BHAGCHEM.NS", "BHAGERIA.NS", "BHAGYANGR.NS",
    "BHANDARI.NS", "BHARATFORG.NS", "BHARATRAS.NS", "BHEL.NS", "BIGBLOC.NS",
    "BIL.NS", "BINANIIND.NS", "BINDALAGRO.NS", "BIRLAMONEY.NS", "BKMINDST.NS",
    "BLISSGVS.NS", "BLS.NS", "BOMDYEING.NS", "BOROLTD.NS", "BPL.NS",
    "BRFL.NS", "BRIGADE.NS", "BRNL.NS", "BROOKS.NS", "BSE.NS",
    "BSHSL.NS", "BSL.NS", "BUTTERFLY.NS", "CADSYS.NS", "CALSOFT.NS",
    "CAMLINFINE.NS", "CAPACITE.NS", "CAPTRUST.NS", "CARERATING.NS", "CARYSIL.NS",
    "CASTROLIND.NS", "CENTENKA.NS", "CENTEXT.NS", "CENTUM.NS", "CERA.NS",
    "CEREBRAINT.NS", "CGCL.NS", "CHALET.NS", "CHEMFAB.NS", "CHEMPLASTS.NS",
    "CHEMTEX.NS", "CHEVIOT.NS", "CHOLAHLDNG.NS", "CIGNITITEC.NS", "CINELINE.NS",
    "CINEVISTA.NS", "CLNINDIA.NS", "COCHINSHIP.NS", "COFFEEDAY.NS", "CONTROLPR.NS",
    "CORALFINAC.NS", "CORDSCABLE.NS", "COROMANDEL.NS", "COSMOFILMS.NS", "COUNCODOS.NS",
    "CREATIVE.NS", "CREST.NS", "CUBEXTUB.NS", "CUPID.NS", "CYBERMEDIA.NS",
    "CYIENT.NS", "DABUR.NS", "DALMIASUG.NS", "DAMODARIND.NS", "DATAMATICS.NS",
    "DBREALTY.NS", "DBSTOCKBRO.NS", "DCAL.NS", "DCMSHRIRAM.NS", "DCW.NS"]

        # --- UI Layout ---
        search_frame = ttk.LabelFrame(master, text="1. Search and Select Tickers", padding="10")
        search_frame.pack(padx=10, pady=5, fill="x")

        ttk.Label(search_frame, text="Search:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.search_entry = AutocompleteCombobox(search_frame, textvariable=self.search_query, width=40)
        self.search_entry.set_completion_list(self.all_tickers)
        self.search_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        self.add_button = ttk.Button(search_frame, text="Add Ticker", command=self.add_ticker)
        self.add_button.grid(row=0, column=2, padx=5, pady=5)

        results_frame = ttk.LabelFrame(master, text="2. Selected Tickers", padding="10")
        results_frame.pack(padx=10, pady=5, fill="both", expand=True)

        self.results_listbox = tk.Listbox(results_frame, selectmode=tk.MULTIPLE, height=10)
        self.results_listbox.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=self.results_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.results_listbox.config(yscrollcommand=scrollbar.set)

        ttk.Label(master, text="Currently Selected:").pack(padx=10, anchor="w")
        self.selected_label = ttk.Label(master, textvariable=self.selected_tickers_display, foreground="darkblue", wraplength=650)
        self.selected_label.pack(padx=10, anchor="w")

        options_frame = ttk.LabelFrame(master, text="3. Data Options", padding="10")
        options_frame.pack(padx=10, pady=5, fill="x")

        ttk.Label(options_frame, text="Period:").grid(row=0, column=0, padx=5, pady=5)
        ttk.Combobox(options_frame, textvariable=self.period_var, values=["1 day", "5 days", "1 month", "3 months", "6 months", "1 year", "2 years", "5 years", "10 years", "YTD", "Max"], state="readonly").grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(options_frame, text="Interval:").grid(row=0, column=2, padx=5, pady=5)
        ttk.Combobox(options_frame, textvariable=self.interval_var, values=["Daily", "Weekly", "Monthly"], state="readonly").grid(row=0, column=3, padx=5, pady=5)

        self.download_button = ttk.Button(master, text="Download to Excel", command=self.download_data)
        self.download_button.pack(padx=10, pady=10, fill="x")

        self.status_label = ttk.Label(master, textvariable=self.status_text, relief=tk.SUNKEN, anchor="w")
        self.status_label.pack(side="bottom", fill="x")
        self.status_text.set("Ready.")

    def add_ticker(self):
        symbol = self.search_query.get().strip()
        if symbol:
            if symbol not in self.search_results_map.values():
                display_text = symbol
                self.search_results_map[display_text] = symbol
                self.results_listbox.insert(tk.END, display_text)
                self.update_selected_tickers_display()
            else:
                messagebox.showinfo("Already Added", f"{symbol} is already in the list.")
        else:
            messagebox.showwarning("Empty Input", "Please enter a ticker.")

    def update_selected_tickers_display(self):
        tickers = [self.results_listbox.get(i) for i in range(self.results_listbox.size())]
        self.selected_tickers_display.set(", ".join(tickers) if tickers else "None")

    def download_data(self):
        selected_indices = self.results_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("No Selection", "Please select one or more tickers.")
            return

        tickers = [self.results_listbox.get(i) for i in selected_indices]

        period_map = {"1 day": "1d", "5 days": "5d", "1 month": "1mo", "3 months": "3mo", "6 months": "6mo", "1 year": "1y", "2 years": "2y", "5 years": "5y", "10 years": "10y", "YTD": "ytd", "Max": "max"}
        interval_map = {"Daily": "1d", "Weekly": "1wk", "Monthly": "1mo"}
        yf_period = period_map.get(self.period_var.get())
        yf_interval = interval_map.get(self.interval_var.get())

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="StockData.xlsx")
        if not file_path:
            return

        writer = pd.ExcelWriter(file_path, engine="openpyxl")
        for ticker in tickers:
            try:
                data = yf.download(ticker, period=yf_period, interval=yf_interval, progress=False)
                if data.empty:
                    continue
                data.to_excel(writer, sheet_name=ticker[:31])
            except Exception as e:
                print(f"Error fetching {ticker}: {e}")

        writer.close()

        # Auto-adjust column width
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:   
                max_length = 0
                column = col[0].column
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[get_column_letter(column)].width = adjusted_width
        wb.save(file_path)

        self.status_text.set(f"Download complete: {os.path.basename(file_path)}")
        messagebox.showinfo("Success", f"Data downloaded to:\n{file_path}")


if __name__ == "__main__":
    root = tk.Tk()
    app = StockDataDownloaderApp(root)
    root.mainloop()
